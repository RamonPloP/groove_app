import math
import os
from dateutil import relativedelta
from flask import Blueprint, request, Response, make_response, render_template, jsonify, send_from_directory, redirect, url_for
from flask_login import login_required, current_user
from io import StringIO
import csv
from sqlalchemy import func, or_, desc
from openpyxl.reader.excel import load_workbook
from sqlalchemy import func
from models.constants import Role
from controllers.studentsController import (index_logic)
from db import db
from models.translates import get_translate
from pytz import timezone
from datetime import datetime
from num2words import num2words

from models.utils import is_admin

employees = Blueprint('employees', __name__)

@employees.before_request
def block_admin_supervisors():
    if current_user.role == Role.ADMIN_SUPERVISOR:
        return redirect(url_for('main.dashboard'))


@employees.route('/employees/all')
@login_required
def index():
    return index_logic()


@employees.route('/employee/show/<employe_id>')
@login_required
def employee_show(employe_id):
    return employee_show_logic(employe_id)


@employees.route('/employee/end/<id>')
@login_required
def employee_end(id):
    employee = Employees.find_by_id(id)
    if current_user.role != 5 and employee.employee_type == 1:
        return render_template('home/page-401.html'), 401
    return employee_end_logic(id)


@employees.route('/employee/blacklist/<id>')
@login_required
def employee_blacklist(id):
    employee = Employees.find_by_id(id)
    if current_user.role != 5 and employee.employee_type == 1:
        return render_template('home/page-401.html'), 401
    return employee_blacklist_logic(id)


@employees.route('/employee/blacklist/crud', methods=['POST'])
@login_required
def employee_blacklist_crud():
    return employee_blacklist_crud_logic()


@employees.route('/employees/addf')
@login_required
def employee_add():
    return employee_add_logic()


@employees.route('/employees/crud', methods=['POST'])
@login_required
def employees_crud():
    return employees_crud_logic()


@employees.route('/employees/crud/add', methods=['POST'])
@login_required
def employees_crud_add():
    return employees_crud_add_logic()


@employees.route('/employees/photo/add', methods=['POST'])
@login_required
def employees_photo_add():
    return employees_photo_add_logic()


@employees.route('/employee/modal/idcard/<employe_id>')
@login_required
def employee_modal_idcard(employe_id):
    return employee_modal_idcard_logic(employe_id)


@employees.route('/employee/crud/end/<employe_id>', methods=['POST'])
@login_required
def employee_crud_end(employe_id):
    return employee_crud_end_logic(employe_id)


@employees.route('/employees/add/pdf/', methods=['post'])
@login_required
def employee_add_pdf():
    return employees_add_pdf()

@employees.route('/employees/modal/addpdf/<employee_id>')
@login_required
def employee_modal_addpdf(employee_id):
    employee = Employees.find_by_id(employee_id)
    if current_user.role != 5 and employee.employee_type == 1:
        return render_template('home/page-401.html'), 401
    return employee_modal_addpdf_logic(employee_id)


@employees.route('/employee/modal/addphoto/<employe_id>')
@login_required
def employee_modal_addphoto(employe_id):
    return employee_modal_addphoto_logic(employe_id)


@employees.route('/employee/vacation/info/<employe_id>')
@login_required
def get_employee_vacation_info(employe_id):
    return Employees.find_by_id_vacation(employe_id)


@employees.route('/employees/export/get', methods=['GET'])
@login_required
def export_data():
    try:
        csv_data = ""
        elements = Employees.export_data()
        if len(elements) > 0:
            for key in elements[0].keys():
                csv_data += f"{get_translate(key)},"
            csv_data += "\n"
            for element in elements:
                for key, value in element.items():
                    if value == None:
                        value = ""
                    csv_data += f"{str(value).replace(',', ' ')},"
                csv_data += "\n"
        csv_data = csv_data.encode("utf_8_sig")
        # Create a direct download response with the CSV data and appropriate headers
        response = Response(csv_data, content_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=empleados.csv"

        return response

    except Exception as e:
        return make_response(e.args, 400)


@employees.route('/employees/report/rh/get', methods=['GET'])
@is_admin
@login_required
def report_rh():
    try:
        csv_data = ""
        elements = Employees.report_rh()
        if len(elements) > 0:
            for key in elements[0].keys():
                csv_data += f"{get_translate(key)},"
            csv_data += "\n"
            for element in elements:
                for key, value in element.items():
                    if value == None:
                        value = ""
                    csv_data += f"{str(value).replace(',', ' ')},"
                csv_data += "\n"
        csv_data = csv_data.encode("utf_8_sig")
        # Create a direct download response with the CSV data and appropriate headers
        response = Response(csv_data, content_type="text/csv")
        response.headers["Content-Disposition"] = "attachment; filename=rh_prenominas_data.csv"

        return response

    except Exception as e:
        return make_response(str(e), 400)


@employees.route('/employees/settlement', methods=['POST'])
@login_required
def employee_settlement():
    data = request.form
    employee_id = data.get('employee_id')
    employee = db.session.query(Employees.__table__,
                                func.concat(Employees.name, " ", Employees.last_name, " ", Employees.middle_name).label(
                                    'full_name'),
                                Banks.name.label("bank_name"), ClientsTradeName.trade_name.label("client_name"),EmployerRegistration.name.label('employer_registration')) \
        .join(Banks, Employees.bank_id == Banks.id, isouter=True) \
        .join(ClientsTradeName, Employees.clients_trade_name_id == ClientsTradeName.id) \
        .join(EmployerRegistration,Employees.employer_registration_id==EmployerRegistration.id)\
        .filter(Employees.id == employee_id).first()
    if employee.status == '1':
        return make_response("El empleado no se encuentra dado de baja aun,favor de darlo de baja primero", 400)
    date_start_end = EmployeesRh.find_by_id_last_complete(employee_id)
    check_settlement = Settlement.find_by_id(employee_id)
    if check_settlement:
        return render_template('home/documents/settlement.html', employee=employee,
                               calculate_data=check_settlement)
    calculate_data = calcutale_settlement(employee, date_start_end, data)
    _settlement = Settlement(
        employee_id = employee_id,
        employee_date_start =  date_start_end.date_start,
        employee_date_end =  date_start_end.date_end,
        day_salary = calculate_data["day_salary"],
        days_passed = calculate_data["days_passed"],
        days_pending_salary = calculate_data["pending_salary"],
        days_until_aguinaldo = calculate_data["days_until_aguinaldo"],
        working_percent = calculate_data["working_percent"],
        days_proportional = calculate_data["days_proportional"],
        aguinaldo_total = calculate_data["aguinaldo_total"],
        days_vacation = calculate_data["days_vacation"],
        vacation_percent = calculate_data["vacation_percent"],
        vacation_proportional = calculate_data["vacation_proportional"],
        vacation_total = calculate_data["vacation_total"],
        vacation_prima = calculate_data["vacation_prima"],
        prima_before = calculate_data["prima_before"],
        pending_days = calculate_data["pending_days"],
        days_20 = 0,
        gratification_90 = calculate_data["gratification_90"],
        debts = calculate_data["debts"],
        total = calculate_data["total"],
        created_by = current_user.id
    )
    db.session.add(_settlement)
    db.session.commit()
    return render_template('home/documents/settlement.html', employee=employee,
                           calculate_data=_settlement)

def calcutale_settlement(employee, date_start_end, extra):
    data = dict()
    data["prima_before"] = 0 if extra.get("prima_before") == '' else float(extra.get("prima_before"))
    data["pending_salary"] = 0 if extra.get("pending_salary") == '' else float(extra.get("pending_salary"))
    data["gratification_90"] = 0 if extra.get("gratification_90") == '' else float(extra.get("gratification_90"))
    data["borrowing"] = 0 if extra.get("borrowing") == '' else float(extra.get("borrowing"))
    data["debts"] = 0 if extra.get("debts") == '' else float(extra.get("debts"))
    fist_day_year = datetime.now().date().replace(month=1, day=1)
    date_start = date_start_end.date_start.date()
    date_end = date_start_end.date_end.date()
    if employee.payroll_id not in [9, 10]:
        payroll = 7
    else:
        payroll = 15
    day_salary = float(employee.salary) / payroll
    data["day_salary"] = day_salary
    data["pending_days"] = data["pending_salary"] * data["day_salary"]
    days_passed = date_start_end.date_end - date_start_end.date_start
    years_passed = relativedelta.relativedelta(date_end, date_start).years

    data["days_passed"] = days_passed.days + 1
    if date_start > fist_day_year:
        operation = date_end - date_start
        data["days_until_aguinaldo"] = operation.days
    else:
        operation = date_end - fist_day_year
        data["days_until_aguinaldo"] = operation.days
    data["working_percent"] = (data["days_until_aguinaldo"] * 100) / 365
    data["days_proportional"] = (data["working_percent"] / 100) * 15
    data["aguinaldo_total"] = data["days_proportional"] * data["day_salary"]
    data["days_vacation"] = vacations(data["days_passed"])
    data["vacation_percent"] = ((data["days_passed"] - (years_passed * 365 + 1) + 1) * 100) / 365
    data["vacation_proportional"] = (data["vacation_percent"] / 100) * data["days_vacation"]
    data["vacation_total"] = data["vacation_proportional"] * data["day_salary"]
    data["vacation_prima"] = data["vacation_total"] * 0.25
    data["total"] = data["vacation_total"] + data["aguinaldo_total"] + data["vacation_prima"] \
                    + data["prima_before"] + data["pending_days"]  + data["gratification_90"] - data[
                        "borrowing"] - data["debts"]
    data["total_text"] = num2words(math.trunc(data["total"]), lang='es', ordinal=False)
    return data


def vacations(days_passed):
    if 1 <= days_passed <= 365:
        return 12
    elif 366 <= days_passed <= 730:
        return 14
    elif 731 <= days_passed <= 1095:
        return 16
    elif 1096 <= days_passed <= 1460:
        return 18
    elif 1461 <= days_passed <= 1825:
        return 20
    elif 1826 <= days_passed <= 3650:
        return 22
    elif 3651 <= days_passed <= 5475:
        return 24
    elif 5476 <= days_passed <= 7300:
        return 26
    elif 7301 <= days_passed <= 9125:
        return 28
    elif 9126 <= days_passed <= 9490:
        return 30
    else:
        return 32


@employees.route('/employee/modal/settlement/<employee_id>', methods=['GET'])
@login_required
def employee_modal_settlement(employee_id):
    employee = Employees.find_by_id(employee_id)
    settlement = Settlement.find_by_id(employee.id)
    return render_template('employees/modal_settlement.html', employee=employee , settlement=settlement)

@employees.route('/employees/modal/signed-settlement/<employee_id>', methods=['GET'])
@login_required
def employee_modal_signed_settlement(employee_id):
    employee = Employees.find_by_id(employee_id)
    settlement = Settlement.find_by_id(employee.id)
    return render_template('employees/modal_add_signed_settlement.html', employee=employee , settlement=settlement)

@employees.route('/employees/add/signed-settlement/', methods=['post'])
@login_required
def employee_add_signed_settlement():
    return employees_add_signed_settlement()

@employees.route('/static/signed_settlements/<string:filename>', methods=['get'])
@login_required
def download_circulation_card_document(filename):
    PATH_FOLDER = os.path.join('static', 'signed_settlements')
    return send_from_directory(PATH_FOLDER, path=filename, as_attachment=True)

@employees.route('/employees/vacation-constancy', methods=['POST'])
@login_required
def employee_vacation_constancy():
    data = request.form
    startDate, endDate = data.get('vacations')[:10], data.get('vacations')[11:]
    startDate = startDate.strip()
    endDate = endDate.strip()

    employee_id = data.get('employee_id')
    employeerh = EmployeesRh.find_by_id_last(employee_id)
    number_of_days = Vacations.find_vacation_by_id(int(data.get("vacation_id"))).number_of_days
    employee = db.session.query(Employees.__table__,
                                func.concat(Employees.name, " ", Employees.last_name, " ", Employees.middle_name).label(
                                    'full_name'),
                                Banks.name.label("bank_name"), ClientsTradeName.trade_name.label("client_name")) \
        .join(Banks, Employees.bank_id == Banks.id, isouter=True) \
        .join(ClientsTradeName, Employees.clients_trade_name_id == ClientsTradeName.id) \
        .filter(Employees.id == employee_id).first()
    date_start_end = EmployeesRh.find_by_id_last_complete(employee_id)
    calculate_data = vacation_constancy(employee, date_start_end, data, startDate, endDate)
    return render_template('home/documents/vacation_constancy.html', employee=employee, date_start_end=date_start_end,
                           calculate_data=calculate_data,number_of_days=number_of_days, employeerh=employeerh)


def vacation_constancy(employee, date_start_end, extra, startDate, endDate):
    data = dict()
    data["vacation_period"] = '' if extra.get("vacation_period") == '' else extra.get("vacation_period")
    startDate = '' if startDate == '' else datetime.strptime(startDate,
                                                                                            '%Y-%m-%d').date()
    endDate = '' if endDate == '' else datetime.strptime(endDate,
                                                                                        '%Y-%m-%d').date()
    data["to_day"] = datetime.now().date()
    data['vacation_start'] = startDate
    data['vacation_end'] = endDate
    date_start = date_start_end.date_start.date()
    data["vacations_days"], data["rsenioity"] = calculate_vacations(date_start)
    if employee.payroll_id not in [9, 10]:
        payroll = 7
    else:
        payroll = 15
    day_salary = float(employee.salary) / payroll
    data["day_salary"] = day_salary
    data["vacation_total"] = data["day_salary"] * 15
    data["vacation_prima"] = data["vacation_total"] * 0.25
    data["total"] = data["vacation_total"] + data["vacation_prima"]
    data["total_text"] = num2words(math.trunc(data["total"]), lang='es', ordinal=False)
    return data


def calculate_vacations(date):
    today = datetime.now()
    delta = relativedelta.relativedelta(today, date)
    years = delta.years
    return {
               1: 12,
               2: 14,
               3: 16,
               4: 18,
               5: 20,
               **dict.fromkeys(range(6, 10), 22),
               **dict.fromkeys(range(11, 15), 24),
               **dict.fromkeys(range(16, 20), 26),
               **dict.fromkeys(range(21, 25), 28),
               **dict.fromkeys(range(26, 30), 30),
               **dict.fromkeys(range(31, 35), 32),
           }.get(years), years


@employees.route('/employee/modal/vacation/<employe_id>', methods=['GET'])
@login_required
def employee_modal_vacation_constancy(employe_id):
    employee = Employees.find_by_id(employe_id)
    return render_template('employees/modal_vacation_constancy.html', employee=employee)

@employees.route('/settlement', methods=['GET'])
@login_required
def settlement():
    settlement = Settlement.get_all()
    return settlement

@employees.route('/employees/settlement/all', methods=['GET'])
@login_required
def employees_settlement_list():
    return render_template('employees/settlement_list.html')

@employees.route('/settlement/<settlement_id>', methods=['GET'])
@login_required
def view_settlement(settlement_id):
    settlement = Settlement.find_by_settlement_id(settlement_id)
    employee = db.session.query(Employees.__table__,
                                func.concat(Employees.name, " ", Employees.last_name, " ", Employees.middle_name).label(
                                    'full_name'),
                                Banks.name.label("bank_name"), ClientsTradeName.trade_name.label("client_name"),
                                EmployerRegistration.name.label('employer_registration')) \
        .join(Banks, Employees.bank_id == Banks.id, isouter=True) \
        .join(ClientsTradeName, Employees.clients_trade_name_id == ClientsTradeName.id) \
        .join(EmployerRegistration, Employees.employer_registration_id == EmployerRegistration.id) \
        .filter(Employees.id == settlement.employee_id).first()
    return render_template('home/documents/settlement.html', employee=employee,
                           calculate_data=settlement)

@employees.route('/employee/modal-edit/settlement/<settlement_id>', methods=['GET'])
@login_required
def employee_modal_edit_settlement(settlement_id):
    settlement = Settlement.find_by_settlement_id(settlement_id)
    return render_template('employees/modal_edit_settlement.html', settlement=settlement)


@employees.route('/employees/settlement/edit', methods=['POST'])
@login_required
def employee_edit_settlement():
    data = request.form
    settlement_id = data.get('settlement_id')
    _settlement = Settlement.find_by_settlement_id(settlement_id)
    employee = db.session.query(Employees.__table__,
                                func.concat(Employees.name, " ", Employees.last_name, " ", Employees.middle_name).label(
                                    'full_name'),
                                Banks.name.label("bank_name"), ClientsTradeName.trade_name.label("client_name"),
                                EmployerRegistration.name.label('employer_registration')) \
        .join(Banks, Employees.bank_id == Banks.id, isouter=True) \
        .join(ClientsTradeName, Employees.clients_trade_name_id == ClientsTradeName.id) \
        .join(EmployerRegistration, Employees.employer_registration_id == EmployerRegistration.id) \
        .filter(Employees.id == _settlement.employee_id).first()
    date_start_end = EmployeesRh.find_by_id_last_complete(employee.id)
    calculate_data = calcutale_settlement(employee, date_start_end, data)
    _settlement.day_salary = calculate_data["day_salary"]
    _settlement.days_passed = calculate_data["days_passed"]
    _settlement.days_pending_salary = calculate_data["pending_salary"]
    _settlement.days_until_aguinaldo = calculate_data["days_until_aguinaldo"]
    _settlement.working_percent = calculate_data["working_percent"]
    _settlement.days_proportional = calculate_data["days_proportional"]
    _settlement.aguinaldo_total = calculate_data["aguinaldo_total"]
    _settlement.days_vacation = calculate_data["days_vacation"]
    _settlement.vacation_percent = calculate_data["vacation_percent"]
    _settlement.vacation_proportional = calculate_data["vacation_proportional"]
    _settlement.vacation_total = calculate_data["vacation_total"]
    _settlement.vacation_prima = calculate_data["vacation_prima"]
    _settlement.prima_before = calculate_data["prima_before"]
    _settlement.pending_days = calculate_data["pending_days"]
    _settlement.days_20 = 0
    _settlement.gratification_90 = calculate_data["gratification_90"]
    _settlement.debts = calculate_data["debts"]
    _settlement.total = calculate_data["total"]
    _settlement.created_by = current_user.id
    db.session.commit()
    return render_template('home/documents/settlement.html', employee=employee,
                           calculate_data=_settlement)

@employees.route('/employee/anniversary', methods=['GET'])
@login_required
def employee_anniversary():
    employees = Employees.company_anniversary()
    return employees

@employees.route('/employees/anniversary/all', methods=['GET'])
@login_required
def employees_anniversary_list():
    return render_template('employees/anniversary_list.html')


@employees.route('/employees/credit/all', methods=['GET'])
@login_required
def employees_credit_list():
    return render_template('employees/credit_list.html')

@employees.route('/employees/credit', methods=['GET'])
@login_required
def get_employee_credit():
    employees = Employees.get_employees_credits()
    return employees

@employees.route('/employee/credit/table/<credit_id>', methods=['GET'])
@login_required
def modal_credit_table(credit_id):
    table = EmployeesCreditPayTable.find_by_credit_id(credit_id)
    return render_template('employees/modal_credit_table.html', table=table)


@employees.route('/employees/update/bulk/',methods=['GET'])
@login_required
def import_employees_view():
    return render_template('employees/bulk_update.html')

@employees.route('/employees/update/bulk/',methods=['POST'])
@login_required
def update_bulk_employees():
    try:
        workbook = load_workbook(request.files['file'])
        # workbook = load_workbook(filename="C:/Users/br_y_/Desktop/clientes.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=3, values_only=False):
            employee = Employees.find_by_curp(row[0].value)
            if not employee:
                return make_response(f'El empleado con la CURP {row[0].value} no esta registrado en el sistema', 400)
            if row[1].value < employee.salary:
                return make_response('No se pueden editar los empleados a un salario menor',400)
            if row[2].value < employee.sdi_imss:
                return make_response('No se pueden editar los empleados a un SDI IMSS menor',400)
            employee.salary = row[1].value
            employee.sdi_imss =  row[2].value
            # update employee
            db.session.commit()


        return make_response('Importacion completada', 201)
    except Exception as e:
        return make_response(str(e), 400)

@employees.route('/employees/report-ups-downs')
@is_admin
@login_required
def ups_downs():
    return render_template('home/reports/download-employees-ups-downs.html')

@employees.route('/employees/report-ups-downs/download-csv', methods=['POST'])
@is_admin
@login_required
def download_ups_downs_csv():
    option = request.form.get('type')
    date_start = request.form.get('date_start')
    date_end = request.form.get('date_end')

    employee = (db.session.query(Employees.nss,Employees.curp, City.name.label('city'),
                                EmployerRegistration.name.label("registro_patronal"),
                                                                func.concat(Employees.last_name, " ", Employees.middle_name, " ",
                                            Employees.name).label('Nombre'),
                                ClientsTradeName.trade_name.label("punto"),
                                EmployeesRh.date_start.label('date_start'),
                                EmployeesRh.date_end.label('date_end'),
                                EmployeesRh.leave_description.label('leave_desc'),
                                EmployeesRh.created_by.label('created_by'),
                                EmployeesRh.removed_by.label('removed_by'),
                                 EmployeesRh.created_date.label('Fecha de registro de alta'),
                                 EmployeesRh.updated_at.label('Fecha de registro de baja'))
                                .join(EmployerRegistration, Employees.employer_registration_id == EmployerRegistration.id)
                                .join(EmployeesRh, Employees.id == EmployeesRh.employee_id)
                                .join(Banks, Employees.bank_id == Banks.id, isouter=True)
                                .join(City, City.id == Employees.location).join(State, State.id == City.state_id)
                                .join(ClientsTradeName, Employees.clients_trade_name_id == ClientsTradeName.id)
                                .join(Clients, ClientsTradeName.client_id == Clients.id))

    if option == 'ups':
        employees = employee.filter(EmployeesRh.date_start.between(date_start, date_end))
    elif option == 'downs':
        employees = employee.filter(EmployeesRh.date_end.between(date_start, date_end))
    elif option == 'both':
        employees = employee.filter(
            or_(
                EmployeesRh.date_start.between(date_start, date_end),
                EmployeesRh.date_end.between(date_start, date_end)
            )
        )
    else:
        return jsonify({'error': 'Invalid option'}), 400

    employees = employees.all()
    elements = [_ele._asdict() for _ele in employees]
    csv_data = ""
    if len(elements) > 0:
        for key in elements[0].keys():
            csv_data += f"{get_translate(key)},"
        csv_data += "\n"
        for element in elements:
            for key, value in element.items():
                if key == 'created_by':
                    value = Users.find_by_id(value)
                    if not value:
                        value = ''
                    else:
                        value = value.name

                if key == 'removed_by':
                    value = Users.find_by_id(value)
                    if not value:
                        value = ''
                    else:
                        value = value.name

                if value == None:
                    value = ""

                if key == 'city':
                    value = value.strip()
                csv_data += f"{str(value).replace(',', ' ')},"
            csv_data += "\n"
    csv_data = csv_data.encode("utf_8_sig")
    # Create a direct download response with the CSV data and appropriate headers
    response = Response(csv_data, content_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=rh_prenominas_data.csv"

    return response


@employees.route('/employees/modal/debts/<employee_id>',methods=['GET'])
@login_required
def debts_employees_modal(employee_id):
    employee = Employees.find_by_id(employee_id)
    if current_user.role != 5 and employee.employee_type == 1:
        return render_template('home/page-401.html'), 401
    return render_template('employees/modal_debts.html',employee=employee)

@employees.route('/employees/debts/all', methods=['GET'])
@login_required
def employees_debts_list():
    return render_template('employees/debts_list.html')


@employees.route('/employees/debts', methods=['GET'])
@login_required
def get_employee_debts():
    employees = Employees.get_employees_debts()
    return employees

@employees.route('/employee/debts/table/<credit_id>', methods=['GET'])
@login_required
def modal_debt_table(credit_id):
    table = EmployeesDebtsPayTable.find_by_credit_id(credit_id)
    return render_template('employees/modal_debts_table.html', table=table)


@employees.route('/employee/debts/add', methods=['POST'])
@login_required
def debt_add():
    data = request.get_json()
    employee_id = data.get('employee_id')
    debt = float(data.get('debt'))
    payment_quantity = float(data.get('payment_quantity'))
    debt_type = data.get('debt_type')
    other_description = data.get('other_description')
    employee = Employees.find_by_id(employee_id)
    try:
        employee_debt = EmployeesDebts(
            employees_id=employee_id,
            debt = debt,
            payment_quantity = payment_quantity,
            debt_type =  debt_type,
            other_description = other_description,
            status=1,
            )
        db.session.add(employee_debt)
        db.session.commit()
        payroll = Payroll.find_by_id(employee.payroll_id)
        if 'JUEVES' in payroll.name:
            weekday = 3
        elif 'MARTES' in payroll.name:
            weekday = 1
        elif 'STAFF' in payroll.name:
            weekday = 5
        last_pay = 0
        pay_days = debt / payment_quantity
        pay = payment_quantity

        if debt % payment_quantity != 0:
            pay_days = int(pay_days) + 1
            last_pay = debt % payment_quantity
        # next_thursday = today + timedelta((3 - today.weekday()) % 7)
        if 'QUINCENAL' in payroll.name:
            next_dates = get_next_byweekly(pay_days)
        else:
            next_dates = get_next_weekday(pay_days, weekday)
        for dates in next_dates:
            if dates == next_dates[-1] and debt % payment_quantity != 0:
                pay = last_pay
            employee_debts_table = EmployeesDebtsPayTable(
                employees_debts_id=employee_debt.id,
                pay_value=pay,
                status=0,
                pay_date=dates
            )
            db.session.add(employee_debts_table)
            db.session.commit()
        return make_response('Registro existoso.', 201)
    except Exception as e:
        return make_response(e.args, 400)
@employees.route('/employees/down-notification/all', methods=['GET'])
@login_required
def employees_down_notification_list():
    return render_template('employees/down_notification_list.html')

@employees.route('/down-notification', methods=['GET'])
@login_required
def employees_down_notification():
    employee_down_notification = EmployeeDownNotification.get_all()
    return employee_down_notification

@employees.route('/down-notification/add/', methods=['POST'])
@login_required
def down_notification_delte():
    data = request.get_json()
    notification_id = data.get('notification_id')
    try:
        employee_down_notification = EmployeeDownNotification.find_by_id(notification_id)
        # insert client
        db.session.delete(employee_down_notification)
        db.session.commit()
        return make_response('borrado existoso.', 201)
    except Exception as e:
        return make_response(e.args, 400)

    
    
@employees.route('/static/employees_certificates/<string:filename>', methods=['get'])
@login_required
def download_pdf_certificate(filename):
    return download_pdf_certificate_logic(filename)


@employees.route('/employees/get/rh/<id>/<type>', methods=['get'])
@login_required
def get_employee_rh(id, type):
    employee = EmployeesRh.find_by_id(id)

    url = employee.link_up_pdf
    return url


@employees.route('/employees/modal/add/inability/<inability_id>',methods=['GET'])
@login_required
def inability_employees_modal(inability_id):
    employee = Employees.find_by_id(inability_id)
    employee_inhability = EmployeesInability.find_last_by_id(inability_id)
    return render_template('employees/modal_add_inability.html',employee=employee, employee_inhability=employee_inhability)

@employees.route('/employees/list-changes/all',methods=['GET'])
@login_required
def list_changes():
    return render_template('employees/list_employee_notifications.html')

@employees.route('/employees/get_changes/',methods=['get'])
@login_required
def get_changes():
    changes = EmployeeChangeNotification.find_all()
    return changes

@employees.route('/employee/changeRegister/',methods=['POST'])
@login_required
def change_register():
    return change_register_employee()

@employees.route('/employee/inability/add', methods=['POST'])
@login_required
def inability_add():
    data = request.get_json()
    employee_id = data.get('employee_id')
    folio = data.get('folio')
    initial_subsequent = data.get('initial_subsequent')
    number_days = data.get('number_days')
    date_start = data.get('date_start')
    date_end = data.get('date_end')
    type = data.get('type')
    employee = Employees.find_by_id(employee_id)
    if initial_subsequent == 'Subsecuente':
        employee_last_inhability = EmployeesInability.find_last_by_id(employee_id)
        if employee_last_inhability:
            employee_last_inhability.status = 0
    try:
        inability = EmployeesInability(
            employees_id=employee_id,
            folio = folio,
            initial_subsequent = initial_subsequent,
            number_days = number_days,
            date_start = date_start,
            date_end = date_end,
            type = type,
            url_inability_photo=None
        )
        db.session.add(inability)
        employee.status = 2
        db.session.commit()
        return make_response('Registro existoso.', 201)
    except Exception as e:
        return make_response(e.args, 400)
@employees.route('/inability', methods=['GET'])
@login_required
def inability():
    inability = EmployeesInability.get_all()
    return inability

@employees.route('/employees/inability/all', methods=['GET'])
@login_required
def employees_inability_list():
    return render_template('employees/inability_list.html')

@employees.route('/employee/change_inhability_status/<id>', methods=['POST'])
@login_required
def change_inhability(id):
    return change_status_inhability(id)


@employees.route('/employees/inability/photo/add/<employee_inability_id>')
@login_required
def employee_modal_inability_addphoto(employee_inability_id):
    employee_inability = EmployeesInability.find_by_id(employee_inability_id)
    employee = Employees.find_by_id(employee_inability.employees_id)

    return render_template('employees/modal_add_inability_photo.html', employee=employee,employee_inability=employee_inability)

@employees.route('/employees/inability/photo/add', methods=['POST'])
@login_required
def employees_photo_inability_add():
    return employees_photo_inability_add_logic()

@employees.route('/employees/modal/change/<employee_id>',methods=['GET'])
@login_required
def change_employees_modal(employee_id):
    employee = Employees.find_by_id(employee_id)
    if current_user.role != 5 and employee.employee_type == 1:
        return render_template('home/page-401.html'), 401
    employerregistration = EmployerRegistration.get_all()
    clients = Clients.get_all()
    return render_template('employees/modal_change_register.html',employee=employee,employerregistration=employerregistration,clients=clients)

@employees.route('/inabilities/list/filters')
@login_required
def inabilities_list_filter():
    return inabilities_list_filter_logic()

@employees.route('/rotations')
@login_required
def report_employee_client():
    report = Employees.employee_rotations_report()
    return render_template('home/reports/rotations.html', report=report)


@employees.route('/rotations/all', methods=['GET'])
@login_required
def rotations_find():
    rotations_list = Employees.employee_rotations_report()
    return rotations_list

@employees.route('/rotations/list', methods=['GET'])
@login_required
def rotations_list():
    return rotations_list

@employees.route('/test')
@login_required
def test():
    return render_template('home/test.html')

@employees.route('/employee/credit/<employee_id>')
def credit_modal(employee_id):
    checks = dict()
    employee = Employees.find_by_id(employee_id)
    if current_user.role != 5 and employee.employee_type == 1:
        return render_template('home/page-401.html'), 401
    date_start = EmployeesRh.find_by_id_last(employee_id)
    check_credits = EmployeesCredit.get_active_employee(employee_id)
    if check_credits:
        checks['active_credit'] = 1
    today = datetime.now()
    delta = relativedelta.relativedelta(today, date_start.date_start)
    if delta.years < 1 and delta.months < 6:
        checks['senioity'] = 1
    return render_template('employees/modal_credit.html', employee_id=employee.id,
                           employee=employee.name + ' ' + employee.last_name + ' ' + employee.middle_name,
                           checks=checks)


@employees.route('/employees/list/advanced-filters')
@login_required
def employees_advanced_list_filter():
    employee_schema = EmployeesSchema(many=True)
    page = request.args.get('page', default=0, type=int)
    draw = request.args.get('draw', 0, int)
    length = request.args.get('length', 10, int)
    start = request.args.get('start', 0, int)
    status=-1
    client=-1
    state=-1
    city_id=-1
    supervisor = request.args.get("supervisor",default=-1, type=int)
    trade_client = request.args.get("client",default=-1, type=int)
    nss = request.args.get("nss",default=-1)
    up_date = request.args.get("up_date",default=-1)
    down_date = request.args.get("down_date",default=-1)

    if up_date == '':
        up_date = -1
    if down_date == '':
        down_date = -1
    if nss == '':
        nss = -1

    search = None
    if 'status' in request.args:
        status = request.args.get('status', default=-1, type=int)
        client = request.args.get('service', default=-1, type=int)
        state = request.args.get('state', default=-1, type=int)
        city_id = request.args.get('city_id', default=-1, type=int)
    if 'search[value]'  in  request.args:
        search = request.args.get('search[value]')
    employees = Employees.find_by_advanced_filters(status,client,state,city_id,length,start,search,supervisor,trade_client,nss,up_date,down_date)
    employees_dump = employee_schema.dump(employees)
    return jsonify(
        draw=draw,
        recordsTotal=employees.total,
        recordsFiltered=employees.total,
        data=employees_dump
    )


@employees.route('/employee/credit/add', methods=['POST'])
@login_required
def credit_add():
    data = request.get_json()
    employee_id = data.get('employee_id')
    credit_value = int(data.get('credit_value'))
    employee = Employees.find_by_id(employee_id)
    if credit_value > 2000:
        return make_response('El monto maximo de credito son $2,000.', 401)
    try:
        employee_credit = EmployeesCredit(
            employees_id=employee_id,
            credit_value=credit_value,
            status=1,
            deposit_date=None
        )
        db.session.add(employee_credit)
        db.session.commit()
        payroll = Payroll.find_by_id(employee.payroll_id)
        if 'JUEVES' in payroll.name:
            weekday = 3
        elif 'MARTES' in payroll.name:
            weekday = 1
        elif 'STAFF' in payroll.name:
            weekday = 5
        last_pay = 0
        if 'QUINCENAL' in payroll.name:
            pay_days = credit_value / CREDIT_PAY_BIWEEKLY
            pay = CREDIT_PAY_BIWEEKLY
        else:
            pay_days = credit_value / CREDIT_PAY_WEEKLY
            pay = CREDIT_PAY_WEEKLY
        if credit_value % CREDIT_PAY_WEEKLY != 0:
            pay_days = int(pay_days) + 1
            last_pay = credit_value % CREDIT_PAY_WEEKLY
        # next_thursday = today + timedelta((3 - today.weekday()) % 7)
        if 'QUINCENAL' in payroll.name:
            next_dates = get_next_byweekly(pay_days)
        else:
            next_dates = get_next_weekday(pay_days, weekday)
        for dates in next_dates:
            if dates == next_dates[-1] and credit_value % CREDIT_PAY_WEEKLY != 0:
                pay = last_pay
            employee_credit_table = EmployeesCreditPayTable(
                employees_credit_id=employee_credit.id,
                pay_value=pay,
                status=0,
                pay_date=dates
            )
            db.session.add(employee_credit_table)
            db.session.commit()
        return make_response('Registro existoso.', 201)
    except Exception as e:
        return make_response(str(e), 400)

@employees.route('/advanced-list')
@login_required
def advanced_list():
    employees = Employees.get_all()
    state = State.get_all()
    services = ClientsTradeName.get_all()
    supervisors = Users.get_all_supervisor_active()
    clientsTradeName = ClientsTradeName.get_all()
    clients = Clients.get_all()
    return render_template('employees/advanced_list.html',clients=clients, trades=clientsTradeName, response=employees, state=state, services=services, supervisors=supervisors)


@employees.route('/employees/letters', methods=['post'])
@login_required
def employee_letters():
    return generate_document()



@employees.route('/employees/modal/letters/<employee_id>', methods=['get'])
@login_required
def employee_modal_letters(employee_id):
    employee = Employees.find_by_id(employee_id)
    if int(employee.status)==1:
        document_type = DOCUMENTS_TYPE.items()
    else:
        document_type = DOCUMENTS_TYPE_INACTIVE.items()
    return render_template('employees/modal_documents.html',document_type=document_type,employee=employee)

@employees.route('/employees/staff-trade/list')
@login_required
def staff_trade_list():
    staff_trade_list = db.session.query(ClientsTradeName.id, ClientsTradeName.trade_name).filter(ClientsTradeName.id.in_([7001, 7002])).all()
    staff_trade_list = [_ele._asdict() for _ele in staff_trade_list]
    return staff_trade_list

@employees.route('/employees/trade-names/list')
@login_required
def trade_list():
    trade_list = db.session.query(ClientsTradeName.id, ClientsTradeName.trade_name, ClientsTradeName.status).filter(ClientsTradeName.id.notin_([7001, 7002])).all()
    trade_list = [_ele._asdict() for _ele in trade_list]
    return trade_list